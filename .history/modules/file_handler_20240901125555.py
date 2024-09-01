import os
import pandas as pd
from urllib.parse import urljoin
import xml.etree.ElementTree as ET
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def load_excel_urls(excel_path):
    df = pd.read_excel(excel_path)
    return df

def parse_sitemap(sitemap_url):
    try:
        response = requests.get(sitemap_url)
        response.raise_for_status()
        
        tree = ET.ElementTree(ET.fromstring(response.content))
        root = tree.getroot()
        namespace = {"ns": "http://www.sitemaps.org/schemas/sitemap/0.9"}

        urls = []
        if root.tag.endswith('sitemapindex'):
            for sitemap in root.findall('ns:sitemap/ns:loc', namespace):
                child_sitemap_url = sitemap.text.strip()
                urls.extend(parse_sitemap(child_sitemap_url))  # Recursively parse child sitemaps
        elif root.tag.endswith('urlset'):
            for url in root.findall('ns:url/ns:loc', namespace):
                urls.append(url.text.strip())
                
        return urls
    except Exception as e:
        print(f"Error parsing sitemap: {e}")
        return []

def save_results_and_report(output_dir, htaccess_results, excel_results, sitemap_results):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Save results to a text file (if needed)
    save_text_results(output_dir, 'htaccess_results.txt', htaccess_results)
    save_text_results(output_dir, 'excel_results.txt', excel_results)
    save_text_results(output_dir, 'sitemap_results.txt', sitemap_results)
    
    # Save results to an Excel file with color formatting
    save_excel_results(output_dir, htaccess_results, excel_results, sitemap_results)

def save_text_results(output_dir, file_name, results):
    output_path = os.path.join(output_dir, file_name)
    with open(output_path, 'w', encoding='utf-8') as file:
        for url, status in results.items():
            file.write(f'{url}: {status}\n')

def save_excel_results(output_dir, htaccess_results, excel_results, sitemap_results):
    output_file = os.path.join(output_dir, 'results.xlsx')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "URL Testergebnisse"
    sheet.append(["URL", "Source", "Result"])

    # Define styles for formatting
    style_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    style_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    style_redirect = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def append_results(results, source):
        for url, status in results.items():
            row = [url, source, status]
            sheet.append(row)
            # Apply formatting based on the status
            row_index = sheet.max_row
            if status == "OK":
                for cell in sheet[row_index]:
                    cell.fill = style_ok
            elif "Redirect" in status:
                for cell in sheet[row_index]:
                    cell.fill = style_redirect
            else:
                for cell in sheet[row_index]:
                    cell.fill = style_error

    append_results(htaccess_results, 'htaccess')
    append_results(excel_results, 'Excel')
    append_results(sitemap_results, 'Sitemap')

    workbook.save(output_file)
    print(f"Results saved to {output_file}")
